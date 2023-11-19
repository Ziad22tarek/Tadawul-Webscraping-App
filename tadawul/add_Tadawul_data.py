import openpyxl as xl
from tadawul.get_Tadawul_data import get_Tadawul_data
from openpyxl.styles import Font
from utilities.get_index_from_list import get_index_from_list
from openpyxl.styles import Alignment




def add_Tadawul_data(file_name):
    '''Add the tadawul data to tadawul data excel


    Parameters
    ------------

    file_name: str
        the name of the report 



    '''


    # get Tadawul data
    data=get_Tadawul_data(file_name)

    # define the tadaul data excel file path
    location=r'.\tadawul data'

    # open Tadawul data excel file
    #wb = xl.load_workbook(location+'\\'+'Tadawul.Trading.Reports.His.xlsx')
    wb = xl.load_workbook(location+'\\'+'Tadawul.Trading.Reports.Data.xlsx')
    ws=wb.get_sheet_by_name(name = 'Tadawul Data') 


    # loop each row in column A
    index_row=[]
    for i in range(1, ws.max_row+1):

        # define emptiness of cell
        if ws.cell(i, 1).value is None:
            # collect indexes of rows
            index_row.append(i)

    # deleting the empty rows 
    if len(index_row)!=0 :
        row_start =index_row[0]
    else:
        row_start =ws.max_row+1


    font_style=Font(name='Calibri',size=9)
    bold_font=Font(bold=True,name='Calibri',size=9)

    percentage_fields_list=data.columns.tolist()
    percentage_fields_indeces=get_index_from_list(percentage_fields_list,'%',last_element=False)
    percentage_list=[]
    for i in percentage_fields_indeces:
        percentage_list.append(percentage_fields_list[i])

    for i in range(len(data)):
        for col in range(len(data.columns)):
            if data.columns[col] in percentage_list:
                ws[chr(65+col)+str(row_start)]=data[data.columns[col]][i]/100
                ws[chr(65+col)+str(row_start)].number_format='0.00%'
            elif data.columns[col]=='Date':
                ws[chr(65+col)+str(row_start)]=data[data.columns[col]][i]
                ws[chr(65+col)+str(row_start)].alignment = Alignment(horizontal='left')
                ws[chr(65+col)+str(row_start)].number_format='d-mmm-yyyy'

            elif data.columns[col] in ['Investor Type','Nationality']:
                ws[chr(65+col)+str(row_start)]=data[data.columns[col]][i]
            else:
                ws[chr(65+col)+str(row_start)]=data[data.columns[col]][i]
                ws[chr(65+col)+str(row_start)].number_format='_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            
            if i==0:
               ws[chr(65+col)+str(row_start)].font=bold_font
            else:
                ws[chr(65+col)+str(row_start)].font=font_style

        row_start+=1


    wb.save(location+'\\'+'Tadawul.Trading.Reports.Data.xlsx') 


    
